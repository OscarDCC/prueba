from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Obtener la fecha y hora actual
fecha_actual = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

# Cargar el archivo Excel existente
workbook = load_workbook('Asistencia.xlsx')

# Obtener la hoja activa (por defecto es la primera hoja)
sheet = workbook.active

# Obtener el número de la última fila ocupada
ultima_fila = sheet.max_row

# Obtener la lista de nombres ya escritos
nombres_escritos = [sheet[f'A{i}'].value for i in range(2, ultima_fila+1)]

# Borrar los datos existentes en la hoja, excepto la primera fila
sheet.delete_rows(2, ultima_fila)

# Obtener la columna correspondiente a la fecha actual
columna_fecha = get_column_letter(3)  # Columna C

# Agregar los encabezados en la primera fila
#sheet['A1'] = 'Nombre'
#sheet['B1'] = 'Asistencia'
#sheet['C1'] = 'Fecha'

# Iterar sobre los nombres y escribir los datos en el archivo
nombres = ['Juan', 'María', 'Pedro', 'María']

for nombre in nombres:
    if nombre not in nombres_escritos:
        nombres_escritos.append(nombre)
        siguiente_fila = sheet.max_row + 1
        sheet[f'A{siguiente_fila}'] = nombre
        sheet[f'B{siguiente_fila}'] = 'Presente'
        sheet[f'C{siguiente_fila}'] = fecha_actual

# Guardar los cambios en el archivo Excel
workbook.save('Asistencia.xlsx')


