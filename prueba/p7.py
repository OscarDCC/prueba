import face_recognition
import cv2
import os
import numpy as np

import subprocess

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Obtener la fecha y hora actual
fecha_actual = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

# Cargar el archivo Excel existente
workbook = load_workbook('Asistencia.xlsx')

# Obtener la hoja activa (por defecto es la primera hoja)
sheet = workbook.active

# Obtener el número de la última fila ocupada
ultima_fila = sheet.max_row

# Borrar los datos existentes en la hoja, excepto la primera fila
sheet.delete_rows(2, ultima_fila)

# Lista de nombres ya escritos
nombres_escritos = []

# Obtener la columna correspondiente a la fecha actual
columna_fecha = get_column_letter(3)  # Columna C

ruta_programa = '/home/jetson/prueba/rele.py'
ruta_programa2 = '/home/jetson/prueba/telearch.py'

KNOWN_FACES_DIR = '/home/jetson/prueba/fotos'
TOLERANCE = 0.6
FRAME_THICKNESS = 3
FONT_THICKNESS = 2
MODEL = 'hog'  # 'hog' or 'cnn'

print('Cargando rostros conocidos...')
known_faces = []
known_names = []

for name in os.listdir(KNOWN_FACES_DIR):
    for filename in os.listdir(f'{KNOWN_FACES_DIR}/{name}'):
        image = face_recognition.load_image_file(f'{KNOWN_FACES_DIR}/{name}/{filename}')
        encoding = face_recognition.face_encodings(image)[0]
        known_faces.append(encoding)
        known_names.append(name)

print('Listo!')
print('Inicializando cámara...')

cap = cv2.VideoCapture(1)

while True:
    ret, frame = cap.read()
    locations = face_recognition.face_locations(frame, model=MODEL)
    encodings = face_recognition.face_encodings(frame, locations)
    for face_encoding, face_location in zip(encodings, locations):
        results = face_recognition.compare_faces(known_faces, face_encoding, TOLERANCE)
        match = None
        if True in results:
            subprocess.run(['python3', ruta_programa])
            match = known_names[results.index(True)]
            print(f' - {match} from {results}')
            top_left = (face_location[3], face_location[0])
            bottom_right = (face_location[1], face_location[2])
            color = [0, 255, 0]
            cv2.rectangle(frame, top_left, bottom_right, color, FRAME_THICKNESS)
            top_left = (face_location[3], face_location[2])
            bottom_right = (face_location[1], face_location[2] + 22)
            cv2.rectangle(frame, top_left, bottom_right, color, cv2.FILLED)
            cv2.putText(frame, match, (face_location[3] + 10, face_location[2] + 15), cv2.FONT_HERSHEY_SIMPLEX, 0.5,
                        (200, 200, 200), FONT_THICKNESS)
            #for match in nombres:
            if match not in nombres_escritos:
                print('Registrado')
                nombres_escritos.append(match)
                siguiente_fila = sheet.max_row + 1
                sheet[f'A{siguiente_fila}'] = match
                sheet[f'B{siguiente_fila}'] = fecha_actual
            workbook.save('Asistencia.xlsx')
	    
        else:
            print(' - desconocido')
            top_left = (face_location[3], face_location[0])
            bottom_right = (face_location[1], face_location[2])
            color = [0, 0, 255]
            cv2.rectangle(frame, top_left, bottom_right, color, FRAME_THICKNESS)

    cv2.imshow('Cámara', frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        subprocess.run(['python3', ruta_programa2])
        break

cap.release()
cv2.destroyAllWindows()


